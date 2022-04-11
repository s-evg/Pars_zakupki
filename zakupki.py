#!/venv/bin/python
"""developer s-evg | https://github.com/s-evg"""

from bs4 import BeautifulSoup as bs
import time
import logging
import openpyxl
from tqdm import tqdm
from appJar import gui
from aiohttp import ClientSession
import asyncio
import pandas as pd
from pathlib import Path
from fake_useragent import UserAgent
import lxml


current_date = time.strftime("%d-%m-%Y—%H-%M")


logging.basicConfig(
    level=logging.DEBUG,
    filename='INFO.log',
    filemode='w',
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S'
)
logging.debug("ВНИМАНИЕ")
logging.info("ОК")
logging.error("ОШИБКА")


URL = "https://zakupki.gov.ru/epz/contract/search/results.html"

HEADERS = {'user-agent': UserAgent().chrome,
           'accept': '*/*'
           }

PARAMS = {
    "morphology": "on",
    "search-filter": "Дате+размещения",
    "fz44=on&contractStageList_0": "on",
    "contractStageList_1": "on",
    "contractStageList_2": "on",
    "contractStageList_3": "on",
    "contractStageList": "0%2C1%2C2%2C3",
    "contractCurrencyID": "-1",
    "budgetLevelsIdNameHidden": "%7B%7D",
    "publishDateFrom": "19.07.2021",
    "sortBy": "UPDATE_DATE",
    "pageNumber": "1",
    "sortDirection": "false",
    "recordsPerPage": "_50",
    "showLotsInfoHidden": "false"
}


#############################################################################################
####          Работа с исходным файлом: проверка и получение номеров аукционов           ####
#############################################################################################

file_extensions = [".xls", ".xlsx", ".xlsm", ".xlsb", ".odf", ".ods", ".odt", ".csv"]

fileTypes = [('all files', '.*'), ('text files', '.txt')]


def check_file(src_file):
    """Проверяем выбран ли файл, и его расширение"""
    errors = False
    errors_msgs = []
    extension = None

    # проверяем что выбран файл
    if src_file == "":
        errors = False
        # если файл не выбран, то собирается новая выгрузка
        # errors_msgs.append("Не выбран файл.")
    else:
        # проверяем что расширение поддерживается
        extension = Path(src_file).suffix.lower()
        if extension not in file_extensions:
            errors = True
            errors_msgs.append("Выбран не поддерживаемый файл.")

    return(errors, errors_msgs, extension)


def read_file(src_file, extension):
    """Открываем файл и получаем множество номеров аукционов, E-mail"""

    if src_file == "":
        return([], [])

    elif extension == ".csv":
        try:
            data_frame = pd.read_csv(src_file, sep=";", on_bad_lines="warn", encoding="utf-8", engine="python")

        except UnicodeDecodeError:
            data_frame = pd.read_csv(src_file, sep=";", on_bad_lines="warn", encoding="cp1252", engine="python")

    else:
        data_frame = pd.read_excel(src_file)

    column_names = data_frame.columns
    number = data_frame[column_names[0]].tolist()
    e_mails = data_frame[column_names[3]].tolist()
    if "№" in str(number[0]):
        number = [str(_.split()[-1]) for _ in number]
    else:
        number = [str(_) for _ in number]

    return(number, e_mails, data_frame)


#############################################################################################
####                 Подготовка задач и сбор информации о поставщиках                    ####
#############################################################################################


info_table = []


async def as_info(sem, session, link, auction, retry=2):
    """Получаем информацию поставщиков по множеству номеров аукционов"""

    try:
        async with session.get(url=link, headers=HEADERS, timeout=60) as response:
            if response.status == 200:
                response_text = await response.text()
                time.sleep(0.03)
            elif response.status == 404:
                info_table.append({
                    "number": auction,
                    "name": "Страница отсутствует",
                    "inn": "Страница отсутствует",
                    "e-mail": "Страница отсутствует",
                    "phone": "Страница отсутствует"
                })
                return False
            else:
                print(f"Сайт не доступен. Ошибка «{response.status}» Проверьте подключение к интернету.")
                print(f"Переподключение через 30 секунд.\nСтраница {link} | Попыток=> {retry}")
                time.sleep(30)
                return await as_info(sem, session, link, auction, retry=(retry - 1))

    except Exception as ex:
        if retry:
            print(f"ОЙ!! Ошибка соединения. Переподключение через 30 секунд.\nСтраница {link} | Попыток=> {retry}")
            print(ex)
            time.sleep(30)
            return await as_info(sem, session, link, auction, retry=(retry - 1))
        else:
            info_table.append({
                "number": auction,
                "name": "ОШИБКА",
                "inn": "ОШИБКА",
                "e-mail": "ОШИБКА",
                "phone": "ОШИБКА"
            })
            return False

    else:
        return response_text


async def semaphore_info(sem, session, link, auction):
    async with sem:
        return await as_info(sem, session, link, auction)


async def gather_data(src_file, extension):
    """Подготавливаем задачи для запроса информации поставщиков"""

    url = "https://zakupki.gov.ru/epz/contract/contractCard/common-info.html?reestrNumber="

    async with ClientSession() as session:
        auctions = await as_reestr_numbers(session)

        tasks = []
        sem = asyncio.Semaphore(5)
        email_check = app.getCheckBox("Добавить только аукционы, по E-mail отсутствующим в базе.")

        auction_source, e_mails_source, date_frame = read_file(src_file, extension)
        print(f"Собрано с сайта: {len(auctions)}\nВ базе: {len(auction_source)}")
        auction_site_set = set(auctions)
        auction_source_set = set(auction_source)
        e_mails_source = set(e_mails_source)
        print(f"Уникально собрано с сайта: {len(auction_site_set)}")

        if email_check:
            print(f"Включена фильтрация по E-mail, будут проверены все {len(auction_site_set)} аукционов.")
        else:
            auctions = auction_site_set - auction_source_set
            print(f"Новых аукционов: {len(auctions)}")

        for auction in auctions:
            link = f"{url}{auction}"
            task = asyncio.create_task(semaphore_info(sem, session, link, auction))
            tasks.append(task)

        print("\nСобираю информацию о поставщиках:")
        info_providers = [await info for info in tqdm(asyncio.as_completed(tasks), total=len(tasks))]

        print("\nРаспарсиваю информацию:")
        for info in tqdm(info_providers):

            if info:

                soup = bs(info, "lxml")
                num = soup.find("div", {"class": "navBreadcrumb__item navBreadcrumb__item_active"}).get_text().split()[1]
                providers = soup.findAll("div", {"class": "participantsInnerHtml"})

                if len(providers) == 0:
                    continue
                else:
                    for provider in providers:
                        if "Телефон, электронная почта" in provider.get_text():
                            td = provider.findAll("td")
                            email = td[-2].get_text().split("\n")[4].strip()
                            # Проверяем включена ли опция проверки e-mail
                            if email_check:
                                # Есть ли полученный e-mail базе, если нет то добавляем этот аукцион
                                if email in e_mails_source:
                                    continue
                            name = td[0].get_text().split("\n")[1].strip()
                            phone = td[-2].get_text().split("\n")[3].strip()
                            inn = td[0].get_text().split("\n\n")
                            for _ in inn:
                                if "инн" in _.lower():
                                    inn = _.split("\n")[-1]

                            info_table.append({
                                "number": num,
                                "name": name,
                                "inn": inn,
                                "e-mail": email,
                                "phone": phone
                            })


#############################################################################################
####                  Подготовка и сбор номеров аукционов с сайта                        ####
#############################################################################################


async def get_page_data(sem, session, pageNumber, retry=2):
    """Собираем страницы аукционов"""
    PARAMS.update(pageNumber=pageNumber)

    try:
        async with session.get(url=URL, headers=HEADERS, params=PARAMS, timeout=60) as response:
            if response.status == 200:
                response_text = await response.text()
                time.sleep(0.03)
            else:
                print(f"Сайт не доступен. Ошибка «{response.status}» Проверьте подключение к интернету.")
                print(f"Переподключение через 20 секунд.\nСтраница {pageNumber} | Попыток=> {retry}")
                time.sleep(20)
                return await get_page_data(sem, session, pageNumber, retry=(retry - 1))

    except Exception as ex:
        if retry:
            print(f"ОЙ!! Ошибка соединения. Переподключение через 20 секунд.\nСтраница {pageNumber} | Попыток=> {retry}")
            print(ex)
            time.sleep(20)
            return await get_page_data(sem, session, pageNumber, retry=(retry - 1))
        else:
            raise

    else:
        return response_text


async def semaphore_age_number(sem, session, pageNumber):
    async with sem:
        return await get_page_data(sem, session, pageNumber)


async def as_reestr_numbers(session):
    """Получаем номера аукционов со страницы с сайта"""

    start = time.time()
    tasks = []
    number_list = []

    sem = asyncio.Semaphore(3)

    for pageNumber in range(1, 101):
        task = asyncio.create_task(semaphore_age_number(sem, session, pageNumber))
        tasks.append(task)

    print("Готовлю список аукционов:")
    page_number = [await pn for pn in tqdm(asyncio.as_completed(tasks), total=len(tasks))]

    print("\nПолучаю номера аукционов:")
    for pn in tqdm(page_number):
        if pn is not None:
            soup = bs(pn, "lxml")
            number = soup.findAll("div", {"class": "registry-entry__header-mid__number"})
            for _ in number:
                link = _.find("a").get("href")
                reestr_number = link.split("=")[-1]
                number_list.append(reestr_number)
        else:
            number_list = []
        time.sleep(0.1)

    print(f"Собрано {len(number_list)}  ссылок за {round((time.time() - start), 2)} секунд")
    return number_list


#############################################################################################
####                            Запись результатов в файл                                ####
#############################################################################################


def write(save_file):
    """Записываем полученные данные в xlsx"""

    book = openpyxl.Workbook()
    sheet = book.active

    sheet['A1'] = 'Номер аукциона'
    sheet['B1'] = 'Наименование компании'
    sheet['C1'] = 'ИНН'
    sheet['D1'] = 'E-mail'
    sheet['E1'] = 'Телефон'

    row = 2

    for _ in info_table:
        sheet[row][0].value = _["number"]
        sheet[row][1].value = _["name"]
        sheet[row][2].value = _["inn"]
        sheet[row][3].value = _["e-mail"]
        sheet[row][4].value = _["phone"]
        row += 1

    while True:

        try:
            if save_file == "":
                if app.getCheckBox("Добавить к имени файла текущую дату и время."):
                    book.save(f'new-zakupki-{current_date}.xlsx')
                else:
                    book.save(f'new-zakupki.xlsx')
            else:
                save_file = save_file.split(".xlsx")[0]
                if app.getCheckBox("Добавить к имени файла текущую дату и время."):
                    book.save(f'{save_file}-{current_date}.xlsx')
                else:
                    book.save(f'{save_file}.xlsx')
            book.close()
            break

        except PermissionError:
            print('Кажется у Вас открыт файл в Exel,\nПожалуйста, закройте его, и тогда я смогу сохранить.')
            input('Закройте Exel, затем вернитесь сюда и нажмите ENTER')


#
#
#


def add_write(src_file, save_file):
    """Создаём новый файл xlsx записав в него новые полученные данные"""

    book = openpyxl.load_workbook(src_file)
    sheet = book.active


    # book = openpyxl.Workbook()
    # sheet = book.active

    sheet['A1'] = 'Номер аукциона'
    sheet['B1'] = 'Наименование компании'
    sheet['C1'] = 'ИНН'
    sheet['D1'] = 'E-mail'
    sheet['E1'] = 'Телефон'

    row = 2

    for _ in info_table:
        sheet.append([
            _["number"],
            _["name"],
            _["inn"],
            _["e-mail"],
            _["phone"],
        ])
        row += 1

    while True:

        try:
            if save_file == "":
                if app.getCheckBox("Добавить к имени файла текущую дату и время."):
                    book.save(f'new-zakupki-{current_date}.xlsx')
                else:
                    book.save(f'new-zakupki.xlsx')
            else:
                save_file = save_file.split(".xlsx")[0]
                if app.getCheckBox("Добавить к имени файла текущую дату и время."):
                    book.save(f'{save_file}-{current_date}.xlsx')
                else:
                    book.save(f'{save_file}.xlsx')
            book.close()
            break

        except PermissionError:
            print('Кажется у Вас открыт файл в Exel,\nПожалуйста, закройте его, и тогда я смогу сохранить.')
            input('Закройте Exel, затем вернитесь сюда и нажмите ENTER')


#############################################################################################
####                            Обработка нажатий кнопок                                 ####
#############################################################################################


def press(button):
    """ Выполняет нажатие кнопки

    Аргументы:
        button: название кнопки. Используем названия Выполнить или Выход
    """
    start = time.time()

    if button == "Старт":
        src_file = app.getEntry("Input_File")
        errors, error_msg, extension = check_file(src_file)
        if errors:
            app.errorBox("Ошибка", "\n".join(error_msg), parent=None)
        else:
            loop = asyncio.get_event_loop()
            loop.run_until_complete(gather_data(src_file, extension))

            save_file = app.getEntry("Output_file")

            # Создать новый файл, или дозаписать инфу в базу, создав новый файл
            if app.getCheckBox("Создать новую базу, добавив в неё новую информацию (доступно только для .xlsx)"):
                add_write(src_file, save_file)
            else:
                write(save_file)

            # print(f"Выполнено за {round((time.time() - start), 2)} секунд")
            stop = time.time() - start
            stop = time.gmtime(stop)
            print(f"Выполнено за {time.strftime('%H:%M:%S', stop)}")
            print("Программу можно закрыть.")

    else:
        app.stop()


def open_file():
    name_file = app.openBox(title="Открыть файл")
    print(name_file)


#############################################################################################
####                  Создать окно пользовательского интерфейса                          ####
#############################################################################################

app = gui(f"Закупки | {current_date}", useTtk=True)
# app.icon = ("icon.ico")
app.setTtkTheme("alt")
app.setSize(500, 200)

# Добавить интерактивные компоненты
app.addLabel("Выберите исходную базу. | Оставьте пустым для новой выгрузки.")
app.addFileEntry("Input_File")

app.addLabel("Выберите файл для сохранения")
# app.addDirectoryEntry("Output_file")  # TODO заменил на выбор файла
app.addSaveEntry("Output_file")

app.addCheckBox("Добавить к имени файла текущую дату и время.")
app.addCheckBox("Добавить только аукционы, по E-mail отсутствующим в базе.")
app.addCheckBox("Создать новую базу, добавив в неё новую информацию (доступно только для .xlsx)")

# Связать кнопки с функцией под названием press
app.addButtons(["Старт"], press)


if __name__ == "__main__":
    app.go()
